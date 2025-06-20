from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    wb = Workbook()
    ws = wb.active
    ws.title = '견적서'

    # 스타일 정의
    header_fill = PatternFill('solid', fgColor='1e3c72')
    header_font = Font(size=20, bold=True, color='FFFFFF')
    section_title_font = Font(size=12, bold=True, color='1e3c72')
    box_fill = PatternFill('solid', fgColor='f8f9fa')
    box_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    table_header_fill = PatternFill('solid', fgColor='1e3c72')
    table_header_font = Font(bold=True, color='FFFFFF')
    table_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    footer_fill = PatternFill('solid', fgColor='1e3c72')
    footer_font = Font(size=10, color='FFFFFF')

    # 상단 로고/타이틀
    ws.merge_cells('A1:B3')
    ws['A1'] = '(로고)'
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('C1:F3')
    ws['C1'] = 'InnoBridge SOLUTION\n(비즈니스의 다리, 최고의 IT 파트너)\n견적서'
    ws['C1'].fill = header_fill
    ws['C1'].font = header_font
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 고객 정보 박스
    ws.merge_cells('A5:F7')
    ws['A5'] = '고객 정보'
    ws['A5'].font = section_title_font
    ws['A5'].fill = box_fill
    ws['A5'].border = box_border
    ws['A5'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A6'] = '회사명'; ws['B6'] = '{{고객회사}}'; ws['C6'] = '담당자'; ws['D6'] = '{{고객담당자}}'; ws['E6'] = '연락처'; ws['F6'] = '{{고객연락처}}'
    ws['A7'] = '이메일'; ws['B7'] = '{{고객이메일}}'
    for col in 'ABCDEF':
        ws[f'{col}6'].border = box_border
        ws[f'{col}7'].border = box_border
        ws[f'{col}6'].fill = box_fill
        ws[f'{col}7'].fill = box_fill

    # 견적 품목 표
    ws.merge_cells('A9:F9')
    ws['A9'] = '견적 품목'
    ws['A9'].font = section_title_font
    ws['A9'].alignment = Alignment(horizontal='left')
    ws.append(['No.', '품목명', '사양/설명', '수량', '단가', '금액'])
    for i, col in enumerate('ABCDEF', 1):
        ws[f'{col}10'].font = table_header_font
        ws[f'{col}10'].fill = table_header_fill
        ws[f'{col}10'].alignment = Alignment(horizontal='center')
        ws[f'{col}10'].border = table_border
    for row in range(11, 16):
        ws.append([row-10, f'{{{{품목{row-10}}}}}', f'{{{{사양{row-10}}}}}', f'{{{{수량{row-10}}}}}', f'{{{{단가{row-10}}}}}', f'{{{{금액{row-10}}}}}'])
        for col in 'ABCDEF':
            ws[f'{col}{row}'].border = table_border

    # 소계/부가세/총금액 박스
    ws.merge_cells('A17:D17')
    ws['A17'] = '소계'
    ws['A17'].alignment = Alignment(horizontal='right')
    ws['E17'] = '{{소계}}'
    ws.merge_cells('A18:D18')
    ws['A18'] = '부가세(10%)'
    ws['A18'].alignment = Alignment(horizontal='right')
    ws['E18'] = '{{부가세}}'
    ws.merge_cells('A19:D19')
    ws['A19'] = '총 금액'
    ws['A19'].alignment = Alignment(horizontal='right')
    ws['E19'] = '{{총금액}}'
    for row in range(17, 20):
        for col in 'ABCDE':
            ws[f'{col}{row}'].border = box_border
        ws[f'{col}{row}'].fill = box_fill

    # 특약 조건 및 유의사항
    ws.merge_cells('A21:F23')
    ws['A21'] = '특약 조건 및 유의사항:\n- 납기: 계약 후 2주 이내 납품\n- 결제 조건: 납품 후 1주일 이내 현금 결제\n- 기타: 문의사항은 담당자에게 연락 바랍니다.'
    ws['A21'].font = Font(size=10)
    ws['A21'].fill = box_fill
    ws['A21'].border = box_border
    ws['A21'].alignment = Alignment(wrap_text=True, vertical='top')

    # 견적 승인/서명란
    ws.merge_cells('A25:C27')
    ws['A25'] = '견적 승인(서명란)'
    ws['A25'].border = box_border
    ws['A25'].alignment = Alignment(horizontal='center', vertical='center')

    # 하단 회사 정보/연락처
    ws.merge_cells('A30:F31')
    ws['A30'] = 'I.B.S InnoBridge SOLUTION | 서울특별시 강남구 ... | Tel: 02-1234-5678 | 담당자: info@ibs-info.com'
    ws['A30'].fill = footer_fill
    ws['A30'].font = footer_font
    ws['A30'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb.save(r'C:\Users\juny3\테스트_엑셀.xlsx')

    with open(r'C:\Users\juny3\테스트_텍스트.txt', 'w', encoding='utf-8') as f:
        f.write('파일 생성 테스트')
except Exception as e:
    with open(r'C:\Users\juny3\텍스트_에러로그.txt', 'w', encoding='utf-8') as f:
        f.write(str(e)) 